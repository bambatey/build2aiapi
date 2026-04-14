"""Analiz sonuçlarını .xlsx olarak dışa aktar — openpyxl.

Her sheet bir tablo: özet, yer değiştirmeler, reaksiyonlar, modlar.
Sayısal hücreler için `number_format` ayarlıdır ki Excel açıldığında
okunaklı gelsin. Dosyayı bellekte oluşturur; router bytes döndürür.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center")


def analysis_to_xlsx(record: dict[str, Any]) -> bytes:
    """Tüm sonucu tek workbook olarak dön."""
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, record)
    _write_modes_sheet(wb, record.get("modes") or [])

    cases = record.get("cases") or {}
    for case_id, case_data in sorted(cases.items()):
        _write_displacements_sheet(wb, case_id, case_data.get("displacements", []))
        _write_reactions_sheet(wb, case_id, case_data.get("reactions", []))

    return _workbook_bytes(wb)


def displacements_to_xlsx(
    case_id: str, displacements: list[dict]
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _write_displacements_sheet(wb, case_id, displacements)
    return _workbook_bytes(wb)


def reactions_to_xlsx(case_id: str, reactions: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _write_reactions_sheet(wb, case_id, reactions)
    return _workbook_bytes(wb)


def modes_to_xlsx(modes: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _write_modes_sheet(wb, modes)
    return _workbook_bytes(wb)


# ---------------------------------------------------------- sheet writers
def _write_summary_sheet(wb: Workbook, record: dict[str, Any]) -> None:
    ws = wb.create_sheet("Özet")
    summary = record.get("summary") or {}
    rows = [
        ("Analiz ID", record.get("analysis_id")),
        ("Durum", record.get("status")),
        ("Süre (ms)", record.get("duration_ms")),
        ("Düğüm sayısı", summary.get("n_nodes")),
        ("Frame eleman", summary.get("n_frame_elements")),
        ("Shell eleman", summary.get("n_shell_elements")),
        ("Serbestlik (free/total)", f"{summary.get('n_dofs_free')} / {summary.get('n_dofs_total')}"),
        ("Yük durumu sayısı", summary.get("n_load_cases")),
        ("Kombinasyon sayısı", summary.get("n_combinations")),
        ("Mod sayısı", summary.get("n_modes")),
        ("Temel periyot (s)", summary.get("fundamental_period")),
        ("Max yer değiştirme (m)", summary.get("max_displacement")),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28


def _write_displacements_sheet(
    wb: Workbook, case_id: str, displacements: list[dict]
) -> None:
    sheet = _safe_sheet_name(f"YD_{case_id}")
    ws = wb.create_sheet(sheet)
    headers = ["Düğüm", "ux (m)", "uy (m)", "uz (m)", "rx (rad)", "ry (rad)", "rz (rad)"]
    _apply_header(ws, headers)
    for r, d in enumerate(displacements, start=2):
        ws.cell(row=r, column=1, value=d.get("node_id"))
        for c, key in enumerate(("ux", "uy", "uz", "rx", "ry", "rz"), start=2):
            cell = ws.cell(row=r, column=c, value=d.get(key))
            cell.number_format = "0.0000E+00"
    _autosize(ws, len(headers))


def _write_reactions_sheet(
    wb: Workbook, case_id: str, reactions: list[dict]
) -> None:
    sheet = _safe_sheet_name(f"RK_{case_id}")
    ws = wb.create_sheet(sheet)
    headers = ["Mesnet", "Fx (kN)", "Fy (kN)", "Fz (kN)",
               "Mx (kN·m)", "My (kN·m)", "Mz (kN·m)"]
    _apply_header(ws, headers)
    for r, d in enumerate(reactions, start=2):
        ws.cell(row=r, column=1, value=d.get("node_id"))
        for c, key in enumerate(("fx", "fy", "fz", "mx", "my", "mz"), start=2):
            cell = ws.cell(row=r, column=c, value=d.get(key))
            cell.number_format = "#,##0.00"
    _autosize(ws, len(headers))


def _write_modes_sheet(wb: Workbook, modes: list[dict]) -> None:
    ws = wb.create_sheet("Modlar")
    headers = [
        "Mod", "T (s)", "f (Hz)", "ω (rad/s)",
        "Mx (%)", "My (%)", "Mz (%)",
    ]
    _apply_header(ws, headers)
    for r, m in enumerate(modes, start=2):
        ws.cell(row=r, column=1, value=m.get("mode_no"))
        ws.cell(row=r, column=2, value=m.get("period")).number_format = "0.0000"
        ws.cell(row=r, column=3, value=m.get("frequency")).number_format = "0.000"
        ws.cell(row=r, column=4, value=m.get("angular_frequency")).number_format = "0.000"
        mp = m.get("mass_participation") or {}
        ws.cell(row=r, column=5, value=(mp.get("ux") or 0) * 100).number_format = "0.00"
        ws.cell(row=r, column=6, value=(mp.get("uy") or 0) * 100).number_format = "0.00"
        ws.cell(row=r, column=7, value=(mp.get("uz") or 0) * 100).number_format = "0.00"
    _autosize(ws, len(headers))


# ----------------------------------------------------------- helpers
def _apply_header(ws, headers: list[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    ws.freeze_panes = "A2"


def _autosize(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[letter].width = 14


def _safe_sheet_name(name: str) -> str:
    """Excel sheet adları max 31 karakter, bazı karakterler yasak."""
    forbidden = set(r"\/:*?[]")
    cleaned = "".join("_" if ch in forbidden else ch for ch in name)
    return cleaned[:31]


def _workbook_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
